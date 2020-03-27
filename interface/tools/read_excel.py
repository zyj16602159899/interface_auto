from openpyxl import load_workbook
import os
from tools.project_path import *
from tools.read_config import ReadConfig

data_file = os.path.join(data_dir,'test_data.xlsx')
config_file = os.path.join(conf_dir,'config.ini')

class ReadExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.case_limits = eval(ReadConfig().read_config('MODE',sheet_name))
    
    def open_file(self):
        self.workbook = load_workbook(self.file_name)
        self.sheet = self.workbook[self.sheet_name]
    
    def close_file(self):
        self.workbook.close()

    def get_data(self):
        self.open_file()
        test_data = []
        if self.case_limits == 'all':
            for i in range(2,self.sheet.max_row+1):
                row_data = {}
                row_data['case'] = self.sheet.cell(i,1).value
                row_data['url'] = self.sheet.cell(i,2).value
                row_data['data'] = self.sheet.cell(i,3).value
                row_data['title'] = self.sheet.cell(i,4).value
                row_data['method'] = self.sheet.cell(i,5).value
                row_data['expected'] = self.sheet.cell(i,6).value
                test_data.append(row_data)
        else:
            for i in self.case_limits:
                row_data = {}
                row_data['case'] = self.sheet.cell(i+1, 1).value
                row_data['url'] = self.sheet.cell(i+1, 2).value
                row_data['data'] = self.sheet.cell(i+1, 3).value
                row_data['title'] = self.sheet.cell(i+1, 4).value
                row_data['method'] = self.sheet.cell(i+1, 5).value
                row_data['expected'] = self.sheet.cell(i+1, 6).value
                test_data.append(row_data)
        self.close_file()
        return test_data

    def write_back_data(self,item,TestResult,return_data):
        self.open_file()
        self.sheet.cell(item,7).value = TestResult
        self.sheet.cell(item,8).value = return_data
        self.workbook.save(self.file_name)
        self.close_file()

if __name__ == '__main__':
    data = ReadExcel(data_file,'recharge').get_data()
    print(len(data))
